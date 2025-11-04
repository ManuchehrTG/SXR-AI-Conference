import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from typing import List

from configs import app_config
from schemes.user import User
from schemes.event import EventStats
from schemes.subscription import SubscriptionStats

# Функция для авто-ширины колонок
def _auto_adjust_columns(worksheet):
	for column_cells in worksheet.columns:
		max_length = 0
		column_letter = get_column_letter(column_cells[0].column)
		
		for cell in column_cells:
			try:
				# Конвертируем значение в строку и получаем длину
				cell_value = str(cell.value) if cell.value is not None else ""
				max_length = max(max_length, len(cell_value))
			except:
				pass
		
		# Добавляем небольшой отступ для лучшего отображения
		adjusted_width = min(max_length + 2, 50)  # Максимальная ширина 50 символов
		worksheet.column_dimensions[column_letter].width = adjusted_width

def save_to_excel(
	user_id: int,
	users: List[User],
	events: List[EventStats],
	subscriptions: List[SubscriptionStats]
) -> str:
	wb_dir = f"{app_config.STORAGE_DIR}/excel"
	os.makedirs(wb_dir, exist_ok=True)

	file_path = f"{wb_dir}/stats_{user_id}.xlsx"

	wb = Workbook()

	ws_users = wb.active
	ws_users.title = "Users"
	ws_users.append([
		"User ID", "Name", "Username", "Email", "Privacy",
		"Studio Coupon Used", "Is admin", "Source", "Created at", "Updated at"
	])

	for user in users:
		ws_users.append([
			user.id, user.first_name, user.username,
			user.email, user.agreed_privacy, user.studio_coupon_used,
			user.is_admin, user.source, user.created_at_str, user.updated_at_str
		])

	ws_events = wb.create_sheet(title="Events")
	ws_events.append(["Event ID", "User ID", "Username", "Type", "Timestamp"])

	for event in events:
		ws_events.append([event.id, event.user_id, event.username, event.type.value, event.ts_str])

	ws_subscriptions = wb.create_sheet(title="Subscriptions")
	ws_subscriptions.append(["Sub ID", "User ID", "Username", "Offer", "Status", "Started at", "Expires at", "Updated at"])

	for sub in subscriptions:
		ws_subscriptions.append([
			sub.id, sub.user_id, sub.username, sub.offer, sub.status, sub.started_at_str, sub.expires_at_str, sub.updated_at_str
		])

	# Применяем авто-ширину к обоим листам
	_auto_adjust_columns(ws_users)
	_auto_adjust_columns(ws_events)
	_auto_adjust_columns(ws_subscriptions)

	wb.save(file_path)

	return file_path
